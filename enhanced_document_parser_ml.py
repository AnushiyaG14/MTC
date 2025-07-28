import openai
import json
import base64
from io import BytesIO
import PyPDF2
import openpyxl
from PIL import Image
import pytesseract
import os
from dotenv import load_dotenv
from pydantic import BaseModel, ValidationError
from typing import List, Dict, Any, Optional
import torch
from transformers import Qwen2VLForConditionalGeneration, AutoTokenizer, AutoProcessor
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, classification_report
import pickle
import numpy as np
from datetime import datetime
import logging
import requests

load_dotenv()

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Enhanced Pydantic models for comprehensive data validation
class DimensionInfo(BaseModel):
    dimension_mm: Optional[float] = 0.0
    section_mm2: Optional[float] = 0.0
    length_mm: Optional[float] = 0.0

class TestingConditions(BaseModel):
    testing_temperature_c: Optional[str] = ""
    brinell_hardness: Optional[str] = ""
    load_kgf: Optional[float] = 0.0
    location: Optional[str] = ""

class ImpactTestInfo(BaseModel):
    impact_test: Optional[str] = ""
    impact_block_size_mm: Optional[str] = ""
    impact_direction: Optional[str] = ""
    energy_value_j_min: Optional[float] = 0.0

class ChemicalComposition(BaseModel):
    C: Optional[float] = 0.0
    Mn: Optional[float] = 0.0
    S: Optional[float] = 0.0
    P: Optional[float] = 0.0
    Si: Optional[float] = 0.0
    Cr: Optional[float] = 0.0
    Ni: Optional[float] = 0.0
    Mo: Optional[float] = 0.0
    Cu: Optional[float] = 0.0
    V: Optional[float] = 0.0

class MechanicalProperties(BaseModel):
    YS_mpa: Optional[float] = 0.0
    UTS_mpa: Optional[float] = 0.0
    EL_percent: Optional[float] = 0.0
    ROA_percent: Optional[float] = 0.0
    hardness_hbw: Optional[str] = ""
    hardness_avg: Optional[float] = 0.0

class ImpactTestResults(BaseModel):
    no_1: Optional[float] = 0.0
    no_2: Optional[float] = 0.0
    no_3: Optional[float] = 0.0
    temperature_c: Optional[str] = ""

class HeatTreatmentProcess(BaseModel):
    process_type: Optional[str] = ""
    temperature_c: Optional[str] = ""
    time_h: Optional[str] = ""
    cooling_method: Optional[str] = ""
    inert_casting: Optional[str] = ""

class HardnessMeasurement(BaseModel):
    no: Optional[int] = 0
    value: Optional[float] = 0.0

class HardnessMeasurements(BaseModel):
    energy_value: Optional[str] = ""
    min_value: Optional[str] = ""
    measurements: List[HardnessMeasurement] = []
    temperature: Optional[str] = ""

class HeatRecord(BaseModel):
    heat_no: Optional[str] = ""
    qty: Optional[int] = 0
    sample: Optional[str] = ""
    chemical_composition: ChemicalComposition = ChemicalComposition()
    mechanical_properties: MechanicalProperties = MechanicalProperties()
    impact_test_results: ImpactTestResults = ImpactTestResults()
    hardness_measurements: HardnessMeasurements = HardnessMeasurements()

class ExtractedData(BaseModel):
    # Header Information
    company_name: Optional[str] = ""
    report_title: Optional[str] = ""
    company_address: Optional[str] = ""
    client: Optional[str] = ""
    po_no: Optional[str] = ""
    po_date: Optional[str] = ""
    date: Optional[str] = ""
    material_specification: Optional[str] = ""
    certificate_type: Optional[str] = ""
    certificate_number: Optional[str] = ""
    
    # Standards and Approvals
    iso_9001_approved: Optional[str] = ""
    ce_2014_68_eu_approved: Optional[str] = ""
    mat_cert_din_standard: Optional[str] = ""
    
    # Component Information
    component: Optional[str] = ""
    name_body: Optional[str] = ""
    
    # Dimensional Information
    tensile_test_bar: DimensionInfo = DimensionInfo()
    
    # Testing Conditions
    testing_conditions: TestingConditions = TestingConditions()
    
    # Impact Test Information
    impact_test_info: ImpactTestInfo = ImpactTestInfo()
    
    # Heat Treatment Process
    heat_treatment: HeatTreatmentProcess = HeatTreatmentProcess()
    
    # Processing Information
    melting_process: Optional[str] = ""
    
    # File Information
    filename: Optional[str] = ""
    upload_date: Optional[str] = ""
    uploaded_by: Optional[str] = ""
    version: Optional[str] = ""
    revision: Optional[str] = ""
    
    # Heat Records
    heat_records: List[HeatRecord] = []

class MLTrainer:
    """Machine Learning trainer for document parsing improvement using Qwen2-VL"""
    
    def __init__(self):
        self.models_dir = "models"
        self.training_data_file = "training_data.json"
        os.makedirs(self.models_dir, exist_ok=True)
        
        # Initialize Qwen2-VL model variables
        self.qwen_model = None
        self.qwen_processor = None
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        
        # Traditional ML models for field classification
        self.field_classifier = None
        self.vectorizer = None
        self.confidence_threshold = 0.8
        
        # Load existing models if available
        self.load_models()
    
    def initialize_qwen_model(self):
        """Initialize Qwen2-VL model for vision-language processing"""
        try:
            model_name = "Qwen/Qwen2-VL-2B-Instruct"
            
            logger.info("Loading Qwen2-VL model...")
            self.qwen_processor = AutoProcessor.from_pretrained(model_name)
            self.qwen_model = Qwen2VLForConditionalGeneration.from_pretrained(
                model_name,
                torch_dtype=torch.float16 if torch.cuda.is_available() else torch.float32,
                device_map="auto" if torch.cuda.is_available() else None
            )
            
            if not torch.cuda.is_available():
                self.qwen_model = self.qwen_model.to(self.device)
            
            logger.info("Qwen2-VL model loaded successfully!")
            return True
            
        except Exception as e:
            logger.error(f"Failed to load Qwen2-VL model: {str(e)}")
            return False
    
    def extract_with_qwen(self, image: Image.Image, extracted_text: str = "") -> Dict[str, Any]:
        """Extract data using Qwen2-VL model for better accuracy"""
        if not self.qwen_model:
            if not self.initialize_qwen_model():
                return {'success': False, 'error': 'Failed to initialize Qwen2-VL model'}
        
        try:
            prompt = """Analyze this heat treatment certificate document and extract ALL visible information in JSON format. 

Extract these categories of information:

1. HEADER INFORMATION:
- company_name, report_title, company_address
- client, po_no, po_date, date
- certificate_number, certificate_type
- iso_9001_approved, ce_2014_68_eu_approved, mat_cert_din_standard

2. MATERIAL & COMPONENT:
- material_specification, component, name_body

3. DIMENSIONAL DATA:
- tensile_test_bar: {dimension_mm, section_mm2, length_mm}

4. TESTING CONDITIONS:
- testing_conditions: {testing_temperature_c, brinell_hardness, load_kgf, location}

5. IMPACT TEST INFO:
- impact_test_info: {impact_test, impact_block_size_mm, impact_direction, energy_value_j_min}

6. HEAT TREATMENT PROCESS:
- heat_treatment: {process_type, temperature_c, time_h, cooling_method, inert_casting}

7. FOR EACH HEAT RECORD:
- heat_no, qty, sample
- chemical_composition: {C, Mn, S, P, Si, Cr, Ni, Mo, Cu, V}
- mechanical_properties: {YS_mpa, UTS_mpa, EL_percent, ROA_percent, hardness_hbw, hardness_avg}
- impact_test_results: {no_1, no_2, no_3, temperature_c}

Return comprehensive JSON with ALL visible fields."""

            messages = [
                {
                    "role": "user",
                    "content": [
                        {"type": "image", "image": image},
                        {"type": "text", "text": prompt}
                    ]
                }
            ]
            
            text_input = self.qwen_processor.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
            inputs = self.qwen_processor(
                text=[text_input],
                images=[image],
                return_tensors="pt"
            ).to(self.device)
            
            with torch.no_grad():
                generated_ids = self.qwen_model.generate(
                    **inputs,
                    max_new_tokens=2048,
                    temperature=0.1,
                    do_sample=False
                )
            
            generated_ids_trimmed = [
                out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
            ]
            
            response_text = self.qwen_processor.batch_decode(
                generated_ids_trimmed, 
                skip_special_tokens=True, 
                clean_up_tokenization_spaces=False
            )[0]
            
            try:
                json_start = response_text.find('{')
                json_end = response_text.rfind('}') + 1
                
                if json_start != -1 and json_end > json_start:
                    json_str = response_text[json_start:json_end]
                    parsed_data = json.loads(json_str)
                    return {'success': True, 'data': parsed_data}
                else:
                    return {'success': False, 'error': 'No valid JSON found in response', 'raw_response': response_text}
                    
            except json.JSONDecodeError as e:
                return {'success': False, 'error': f'JSON parsing failed: {str(e)}', 'raw_response': response_text}
        
        except Exception as e:
            logger.error(f"Qwen2-VL extraction failed: {str(e)}")
            return {'success': False, 'error': f'Qwen2-VL extraction failed: {str(e)}'}
    
    def save_training_sample(self, extracted_text: str, parsed_data: Dict, validation_result: Dict, user_corrections: Dict = None):
        """Save training sample for future model improvement"""
        training_sample = {
            'timestamp': datetime.now().isoformat(),
            'extracted_text': extracted_text,
            'parsed_data': parsed_data,
            'validation_result': validation_result,
            'user_corrections': user_corrections or {},
            'is_valid': validation_result.get('isValid', False)
        }
        
        training_data = self.load_training_data()
        training_data.append(training_sample)
        
        with open(self.training_data_file, 'w') as f:
            json.dump(training_data, f, indent=2)
        
        logger.info(f"Saved training sample. Total samples: {len(training_data)}")
    
    def load_training_data(self) -> List[Dict]:
        """Load existing training data"""
        try:
            if os.path.exists(self.training_data_file):
                with open(self.training_data_file, 'r') as f:
                    return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load training data: {str(e)}")
        return []
    
    def train_field_classifier(self):
        """Train a field classifier using existing data"""
        training_data = self.load_training_data()
        
        if len(training_data) < 10:
            logger.warning("Not enough training data for model training")
            return False
        
        try:
            texts = []
            labels = []
            
            for sample in training_data:
                if sample['is_valid']:
                    text = sample['extracted_text']
                    parsed_data = sample['parsed_data']
                    
                    for field_name, field_value in parsed_data.items():
                        if isinstance(field_value, (str, int, float)) and field_value:
                            texts.append(text)
                            labels.append(field_name)
            
            if len(texts) < 5:
                logger.warning("Not enough valid training examples")
                return False
            
            self.vectorizer = TfidfVectorizer(max_features=1000, stop_words='english')
            X = self.vectorizer.fit_transform(texts)
            
            X_train, X_test, y_train, y_test = train_test_split(X, labels, test_size=0.2, random_state=42)
            
            self.field_classifier = RandomForestClassifier(n_estimators=100, random_state=42)
            self.field_classifier.fit(X_train, y_train)
            
            y_pred = self.field_classifier.predict(X_test)
            accuracy = accuracy_score(y_test, y_pred)
            
            logger.info(f"Field classifier trained with accuracy: {accuracy:.2f}")
            
            self.save_models()
            return True
            
        except Exception as e:
            logger.error(f"Training failed: {str(e)}")
            return False
    
    def save_models(self):
        """Save trained models"""
        try:
            if self.field_classifier and self.vectorizer:
                with open(os.path.join(self.models_dir, 'field_classifier.pkl'), 'wb') as f:
                    pickle.dump(self.field_classifier, f)
                
                with open(os.path.join(self.models_dir, 'vectorizer.pkl'), 'wb') as f:
                    pickle.dump(self.vectorizer, f)
                
                logger.info("Models saved successfully")
        except Exception as e:
            logger.error(f"Failed to save models: {str(e)}")
    
    def load_models(self):
        """Load trained models"""
        try:
            classifier_path = os.path.join(self.models_dir, 'field_classifier.pkl')
            vectorizer_path = os.path.join(self.models_dir, 'vectorizer.pkl')
            
            if os.path.exists(classifier_path) and os.path.exists(vectorizer_path):
                with open(classifier_path, 'rb') as f:
                    self.field_classifier = pickle.load(f)
                
                with open(vectorizer_path, 'rb') as f:
                    self.vectorizer = pickle.load(f)
                
                logger.info("Models loaded successfully")
        except Exception as e:
            logger.error(f"Failed to load models: {str(e)}")
    
    def get_model_confidence(self, text: str) -> float:
        """Get confidence score for field extraction"""
        if not self.field_classifier or not self.vectorizer:
            return 0.5
        
        try:
            X = self.vectorizer.transform([text])
            probabilities = self.field_classifier.predict_proba(X)
            return np.max(probabilities) if len(probabilities) > 0 else 0.5
        except:
            return 0.5

class EnhancedDocumentParser:
    def __init__(self):
        self.openai_client = openai.OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        self.ml_trainer = MLTrainer()
        
    def extract_text_from_file(self, file_content: bytes, file_type: str, filename: str) -> str:
        """Extract text from different file types"""
        try:
            if file_type.startswith('application/pdf'):
                return self._extract_from_pdf(file_content)
            elif file_type.startswith('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') or file_type.startswith('application/vnd.ms-excel'):
                return self._extract_from_excel(file_content)
            elif file_type.startswith('image/'):
                return self._extract_from_image(file_content)
            else:
                return file_content.decode('utf-8', errors='ignore')
        except Exception as e:
            raise Exception(f"Failed to extract text from {filename}: {str(e)}")
    
    def _extract_from_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF"""
        pdf_file = BytesIO(file_content)
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    
    def _extract_from_excel(self, file_content: bytes) -> str:
        """Extract text from Excel file"""
        excel_file = BytesIO(file_content)
        workbook = openpyxl.load_workbook(excel_file)
        text = ""
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n--- Sheet: {sheet_name} ---\n"
            
            for row in sheet.iter_rows(values_only=True):
                row_text = []
                for cell in row:
                    if cell is not None:
                        row_text.append(str(cell))
                if row_text:
                    text += " | ".join(row_text) + "\n"
        
        return text
    
    def _extract_from_image(self, file_content: bytes) -> str:
        """Extract text from image using OCR"""
        try:
            image = Image.open(BytesIO(file_content))
            text = pytesseract.image_to_string(image)
            return text
        except Exception as e:
            return f"OCR extraction failed: {str(e)}"
    
    def parse_with_ml_enhanced(self, file_content: bytes, file_type: str, extracted_text: str, filename: str) -> Dict[str, Any]:
        """Parse document using ML-enhanced approach with Qwen2-VL"""
        
        # For images, use Qwen2-VL for better extraction
        if file_type.startswith('image/'):
            try:
                image = Image.open(BytesIO(file_content))
                qwen_result = self.ml_trainer.extract_with_qwen(image, extracted_text)
                
                if qwen_result['success']:
                    return {
                        'success': True,
                        'data': qwen_result['data'],
                        'method': 'qwen2vl'
                    }
            except Exception as e:
                logger.warning(f"Qwen2-VL extraction failed, falling back to OpenAI: {str(e)}")
        
        # Fallback to enhanced OpenAI parsing
        return self.parse_with_openai_enhanced(extracted_text, filename)
    
    def parse_with_openai_enhanced(self, extracted_text: str, filename: str) -> Dict[str, Any]:
        """Enhanced OpenAI parsing with comprehensive field extraction"""
        
        system_prompt = """You are a specialized document parser for heat treatment certificates. 
        Extract ALL structured data from the provided text and return it in the exact JSON format specified.
        
        Based on the heat treatment certificate format, extract these comprehensive details:

        1. HEADER & COMPANY INFORMATION:
        - company_name, report_title, company_address
        - client, po_no, po_date, date
        - certificate_number, certificate_type
        - iso_9001_approved, ce_2014_68_eu_approved, mat_cert_din_standard

        2. MATERIAL & COMPONENT INFO:
        - material_specification, component, name_body

        3. DIMENSIONAL DATA:
        - tensile_test_bar: {dimension_mm, section_mm2, length_mm}

        4. TESTING CONDITIONS:
        - testing_conditions: {testing_temperature_c, brinell_hardness, load_kgf, location}

        5. IMPACT TEST INFO:
        - impact_test_info: {impact_test, impact_block_size_mm, impact_direction, energy_value_j_min}

        6. HEAT TREATMENT PROCESS:
        - heat_treatment: {process_type, temperature_c, time_h, cooling_method, inert_casting}

        7. FOR EACH HEAT RECORD:
        - heat_no, qty, sample
        - chemical_composition: {C, Mn, S, P, Si, Cr, Ni, Mo, Cu, V}
        - mechanical_properties: {YS_mpa, UTS_mpa, EL_percent, ROA_percent, hardness_hbw, hardness_avg}
        - impact_test_results: {no_1, no_2, no_3, temperature_c}
        - hardness_measurements: {energy_value, min_value, measurements: [{no, value}], temperature}

        Return comprehensive JSON with ALL visible fields. Convert numeric values appropriately."""
        
        try:
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"Parse this heat treatment certificate from file '{filename}':\n\n{extracted_text}"}
                ],
                response_format={"type": "json_object"},
                temperature=0.1
            )
            
            parsed_data = json.loads(response.choices[0].message.content)
            return {
                'success': True,
                'data': parsed_data,
                'method': 'openai'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"OpenAI parsing failed: {str(e)}"
            }
    
    def validate_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Validate extracted data using enhanced Pydantic models"""
        try:
            validated_data = ExtractedData(**data)
            return {
                'isValid': True,
                'data': validated_data.dict(),
                'errors': []
            }
        except ValidationError as e:
            return {
                'isValid': False,
                'data': data,
                'errors': [str(error) for error in e.errors]
            }
    
    def create_separate_records(self, validated_data: Dict[str, Any]) -> Dict[str, Any]:
        """Create separate records for each Heat No with enhanced data"""
        if not validated_data['isValid']:
            return {
                'success': False,
                'error': 'Data validation failed',
                'validationErrors': validated_data['errors']
            }
        
        data = validated_data['data']
        base_data = {k: v for k, v in data.items() if k != 'heat_records'}
        
        separate_files = []
        for heat_record in data['heat_records']:
            record = {
                **base_data,
                'heatRecord': heat_record,
                'filename': f"{base_data.get('filename', 'unknown')}_Heat_{heat_record.get('heat_no', 'unknown')}.json"
            }
            separate_files.append(record)
        
        return {
            'success': True,
            'totalRecords': len(separate_files),
            'records': separate_files
        }
    
    def convert_to_csv_format(self, separate_records: Dict[str, Any]) -> Dict[str, Any]:
        """Convert to comprehensive CSV format with all fields"""
        if not separate_records['success']:
            return {'success': False, 'error': separate_records['error']}
        
        csv_records = []
        for record in separate_records['records']:
            # Flatten all nested structures for CSV
            flat_record = {}
            
            def flatten_dict(d, prefix=''):
                for k, v in d.items():
                    if isinstance(v, dict):
                        flatten_dict(v, f"{prefix}{k}_")
                    elif isinstance(v, list):
                        if v and isinstance(v[0], dict):
                            for i, item in enumerate(v):
                                flatten_dict(item, f"{prefix}{k}_{i}_")
                        else:
                            flat_record[f"{prefix}{k}"] = json.dumps(v) if v else ""
                    else:
                        flat_record[f"{prefix}{k}"] = v
            
            flatten_dict(record)
            csv_records.append(flat_record)
        
        return {
            'success': True,
            'csvData': csv_records
        }
    
    def process_document(self, file_content: bytes, file_type: str, filename: str, enable_training: bool = True) -> Dict[str, Any]:
        """Main method to process document with ML enhancement"""
        try:
            # Step 1: Extract text
            extracted_text = self.extract_text_from_file(file_content, file_type, filename)
            
            # Step 2: Parse with ML-enhanced approach
            parsing_result = self.parse_with_ml_enhanced(file_content, file_type, extracted_text, filename)
            if not parsing_result['success']:
                return parsing_result
            
            # Step 3: Validate data
            validated_data = self.validate_data(parsing_result['data'])
            
            # Step 4: Create separate records
            separate_records = self.create_separate_records(validated_data)
            
            # Step 5: Convert to CSV
            csv_data = self.convert_to_csv_format(separate_records)
            
            # Step 6: Save training sample if enabled
            if enable_training:
                self.ml_trainer.save_training_sample(
                    extracted_text, 
                    parsing_result['data'], 
                    validated_data
                )
            
            # Step 7: Check if we should retrain models
            training_data = self.ml_trainer.load_training_data()
            if len(training_data) % 10 == 0 and len(training_data) > 0:
                logger.info("Retraining models with new data...")
                self.ml_trainer.train_field_classifier()
            
            return {
                'extractedText': extracted_text,
                'structuredData': validated_data,
                'separateRecords': separate_records,
                'csvData': csv_data,
                'summary': {
                    'totalHeatRecords': separate_records['totalRecords'] if separate_records['success'] else 0,
                    'validationStatus': 'PASSED' if validated_data['isValid'] else 'FAILED',
                    'validationErrors': validated_data['errors'],
                    'extractionMethod': parsing_result.get('method', 'unknown'),
                    'modelConfidence': self.ml_trainer.get_model_confidence(extracted_text)
                }
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Document processing failed: {str(e)}"
            }